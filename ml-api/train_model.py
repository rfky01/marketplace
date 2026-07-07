import argparse
import hashlib
import json
import shutil
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path

import joblib
import numpy as np
import sklearn
from openpyxl import load_workbook
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics import (
    accuracy_score,
    classification_report,
    confusion_matrix,
    f1_score,
)
from sklearn.model_selection import GridSearchCV, StratifiedGroupKFold
from sklearn.pipeline import FeatureUnion
from sklearn.tree import DecisionTreeClassifier

from text_preprocessing import preprocessing_teks


BASE_DIR = Path(__file__).resolve().parent
DEFAULT_DATASET = BASE_DIR / "data" / "dataset_produk_umkm.xlsx"
ACTIVE_MODEL = BASE_DIR / "pipeline_decision_tree_umkm.pkl"
REPORT_PATH = BASE_DIR / "training_report.json"
ALLOWED_LABELS = {"makanan", "kerajinan", "pertanian", "perikanan"}

KAMUS_NORMALISASI = {
    "bbit": "bibit",
    "kean": "kesan",
    "kripik": "keripik",
    "kwalitas": "kualitas",
    "lezaat": "lezat",
    "lezaatt": "lezat",
    "lezattt": "lezat",
    "muanis": "manis",
    "poll": "sekali",
    "psang": "pisang",
    "rapih": "rapi",
    "ruenyah": "renyah",
    "ruuuenyah": "renyah",
    "superr": "super",
    "tinggin": "tinggi",
    "toop": "top",
}


def parse_args():
    parser = argparse.ArgumentParser(
        description="Training TF-IDF + Decision Tree untuk klasifikasi produk PangkalMart."
    )
    parser.add_argument("--dataset", type=Path, default=DEFAULT_DATASET)
    parser.add_argument("--sheet", default="Dataset_1000")
    parser.add_argument("--free-test-sheet", default="Uji_Input_Bebas")
    parser.add_argument(
        "--activate",
        action="store_true",
        help="Backup model lama dan aktifkan model baru setelah evaluasi berhasil.",
    )
    parser.add_argument(
        "--min-holdout-f1",
        type=float,
        default=0.90,
        help="Batas minimum macro F1 holdout agar model boleh diaktifkan.",
    )
    return parser.parse_args()


def read_sheet(path: Path, sheet_name: str, expected_columns):
    workbook = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' tidak ditemukan. Tersedia: {workbook.sheetnames}")

    sheet = workbook[sheet_name]
    rows = sheet.iter_rows(values_only=True)
    header = [str(value or "").strip() for value in next(rows)]
    indexes = {}

    for column in expected_columns:
        if column not in header:
            raise ValueError(f"Kolom '{column}' tidak ditemukan pada sheet '{sheet_name}'.")
        indexes[column] = header.index(column)

    records = []
    for row_number, row in enumerate(rows, start=2):
        record = {
            column: str(row[indexes[column]] or "").strip()
            for column in expected_columns
        }
        if not any(record.values()):
            continue
        if not all(record.values()):
            raise ValueError(f"Data kosong ditemukan pada baris {row_number} sheet '{sheet_name}'.")
        records.append(record)

    return records


def load_training_data(path: Path, sheet_name: str):
    records = read_sheet(
        path,
        sheet_name,
        ["nama_produk", "deskripsi_produk", "kategori"],
    )

    labels = [record["kategori"].lower() for record in records]
    invalid_labels = sorted(set(labels) - ALLOWED_LABELS)
    if invalid_labels:
        raise ValueError(f"Kategori tidak valid ditemukan: {invalid_labels}")

    duplicate_count = len(records) - len(
        {
            (
                record["nama_produk"].lower(),
                record["deskripsi_produk"].lower(),
                record["kategori"].lower(),
            )
            for record in records
        }
    )
    if duplicate_count:
        raise ValueError(f"Dataset memiliki {duplicate_count} baris duplikat persis.")

    name_to_labels = {}
    for record in records:
        name = record["nama_produk"].strip().lower()
        name_to_labels.setdefault(name, set()).add(record["kategori"].lower())
    conflicts = {name: values for name, values in name_to_labels.items() if len(values) > 1}
    if conflicts:
        raise ValueError(f"Nama produk memiliki kategori konflik: {list(conflicts.items())[:5]}")

    texts = [
        preprocessing_teks(
            f'{record["nama_produk"]} {record["nama_produk"]} {record["deskripsi_produk"]}',
            KAMUS_NORMALISASI,
        )
        for record in records
    ]
    groups = [record["nama_produk"].strip().lower() for record in records]

    return records, np.array(texts), np.array(labels), np.array(groups)


def load_free_test(path: Path, sheet_name: str):
    records = read_sheet(
        path,
        sheet_name,
        ["nama_produk", "deskripsi_produk", "kategori_seharusnya"],
    )
    texts = [
        preprocessing_teks(
            f'{record["nama_produk"]} {record["nama_produk"]} {record["deskripsi_produk"]}',
            KAMUS_NORMALISASI,
        )
        for record in records
    ]
    labels = [record["kategori_seharusnya"].lower() for record in records]
    return records, texts, labels


def build_vectorizers():
    return {
        "word_tfidf": TfidfVectorizer(
            ngram_range=(1, 3),
            min_df=1,
            max_df=0.98,
            max_features=10000,
            sublinear_tf=True,
        ),
        "hybrid_word_char_tfidf": FeatureUnion(
            [
                (
                    "word",
                    TfidfVectorizer(
                        ngram_range=(1, 3),
                        min_df=1,
                        max_df=0.98,
                        max_features=10000,
                        sublinear_tf=True,
                    ),
                ),
                (
                    "char",
                    TfidfVectorizer(
                        analyzer="char_wb",
                        ngram_range=(3, 5),
                        min_df=2,
                        max_features=12000,
                        sublinear_tf=True,
                    ),
                ),
            ]
        ),
    }


def evaluate_candidate(name, vectorizer, texts, labels, groups, train_idx, test_idx):
    train_texts = texts[train_idx]
    test_texts = texts[test_idx]
    y_train = labels[train_idx]
    y_test = labels[test_idx]
    train_groups = groups[train_idx]

    x_train = vectorizer.fit_transform(train_texts)
    x_test = vectorizer.transform(test_texts)

    inner_cv = StratifiedGroupKFold(n_splits=4, shuffle=True, random_state=43)
    cv_splits = list(inner_cv.split(x_train, y_train, train_groups))

    parameter_grid = {
        "criterion": ["gini", "entropy"],
        "max_depth": [12, 15, 20, None],
        "min_samples_split": [2, 5, 10],
        "min_samples_leaf": [1, 2, 5],
        "class_weight": [None, "balanced"],
    }

    search = GridSearchCV(
        DecisionTreeClassifier(random_state=42),
        parameter_grid,
        scoring="f1_macro",
        cv=cv_splits,
        n_jobs=-1,
        refit=True,
    )
    search.fit(x_train, y_train)

    predictions = search.best_estimator_.predict(x_test)
    report = classification_report(y_test, predictions, output_dict=True, zero_division=0)

    return {
        "name": name,
        "vectorizer": vectorizer,
        "model": search.best_estimator_,
        "cv_macro_f1": float(search.best_score_),
        "holdout_accuracy": float(accuracy_score(y_test, predictions)),
        "holdout_macro_f1": float(f1_score(y_test, predictions, average="macro")),
        "best_params": search.best_params_,
        "classification_report": report,
        "confusion_matrix": confusion_matrix(
            y_test,
            predictions,
            labels=["makanan", "kerajinan", "pertanian", "perikanan"],
        ).tolist(),
    }


def evaluate_free_test(candidate, records, texts, labels):
    x_free = candidate["vectorizer"].transform(texts)
    predictions = candidate["model"].predict(x_free)
    results = []
    for record, expected, predicted in zip(records, labels, predictions):
        results.append(
            {
                "nama_produk": record["nama_produk"],
                "expected": expected,
                "predicted": str(predicted),
                "correct": expected == predicted,
            }
        )
    return float(accuracy_score(labels, predictions)), results


def evaluate_old_model(texts, labels, free_texts, free_labels):
    if not ACTIVE_MODEL.exists():
        return None
    try:
        package = joblib.load(ACTIVE_MODEL)
        old_model = package["model"]
        old_tfidf = package["tfidf"]
        predictions = old_model.predict(old_tfidf.transform(texts))
        free_predictions = old_model.predict(old_tfidf.transform(free_texts))
        return {
            "dataset_accuracy": float(accuracy_score(labels, predictions)),
            "dataset_macro_f1": float(f1_score(labels, predictions, average="macro")),
            "free_test_accuracy": float(accuracy_score(free_labels, free_predictions)),
        }
    except Exception as error:
        return {"error": str(error)}


def dataset_sha256(path: Path):
    digest = hashlib.sha256()
    with path.open("rb") as file:
        for chunk in iter(lambda: file.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def json_safe(value):
    if isinstance(value, (np.integer, np.floating)):
        return value.item()
    if isinstance(value, np.ndarray):
        return value.tolist()
    raise TypeError(f"Tidak dapat mengubah {type(value)} ke JSON")


def main():
    args = parse_args()
    dataset_path = args.dataset.resolve()
    if not dataset_path.exists():
        raise FileNotFoundError(f"Dataset tidak ditemukan: {dataset_path}")

    records, texts, labels, groups = load_training_data(dataset_path, args.sheet)
    free_records, free_texts, free_labels = load_free_test(dataset_path, args.free_test_sheet)

    outer_cv = StratifiedGroupKFold(n_splits=5, shuffle=True, random_state=42)
    train_idx, test_idx = next(outer_cv.split(texts, labels, groups))

    candidates = []
    for name, vectorizer in build_vectorizers().items():
        print(f"Training candidate: {name}")
        candidate = evaluate_candidate(
            name,
            vectorizer,
            texts,
            labels,
            groups,
            train_idx,
            test_idx,
        )
        free_accuracy, free_results = evaluate_free_test(
            candidate,
            free_records,
            free_texts,
            free_labels,
        )
        candidate["free_test_accuracy"] = free_accuracy
        candidate["free_test_results"] = free_results
        candidates.append(candidate)
        print(
            f"  CV macro F1={candidate['cv_macro_f1']:.4f}, "
            f"holdout F1={candidate['holdout_macro_f1']:.4f}, "
            f"free test={free_accuracy:.4f}"
        )

    best = max(
        candidates,
        key=lambda item: (
            item["cv_macro_f1"],
            item["holdout_macro_f1"],
            item["free_test_accuracy"],
        ),
    )

    if best["holdout_macro_f1"] < args.min_holdout_f1:
        raise RuntimeError(
            f"Model tidak diaktifkan: holdout macro F1 {best['holdout_macro_f1']:.4f} "
            f"di bawah batas {args.min_holdout_f1:.4f}."
        )

    final_vectorizer = build_vectorizers()[best["name"]]
    final_x = final_vectorizer.fit_transform(texts)
    final_model = DecisionTreeClassifier(random_state=42, **best["best_params"])
    final_model.fit(final_x, labels)

    trained_at = datetime.now(timezone.utc).isoformat()
    metadata = {
        "algorithm": "TF-IDF + Decision Tree",
        "feature_strategy": best["name"],
        "dataset_file": dataset_path.name,
        "dataset_sheet": args.sheet,
        "dataset_sha256": dataset_sha256(dataset_path),
        "row_count": len(records),
        "unique_product_names": len(set(groups)),
        "class_distribution": dict(Counter(labels)),
        "trained_at_utc": trained_at,
        "sklearn_version": sklearn.__version__,
        "best_params": best["best_params"],
        "cv_macro_f1": best["cv_macro_f1"],
        "holdout_accuracy": best["holdout_accuracy"],
        "holdout_macro_f1": best["holdout_macro_f1"],
        "free_test_accuracy": best["free_test_accuracy"],
    }

    package = {
        "model": final_model,
        "tfidf": final_vectorizer,
        "kamus_normalisasi": KAMUS_NORMALISASI,
        "metadata": metadata,
    }

    candidate_path = BASE_DIR / "pipeline_decision_tree_umkm_candidate.pkl"
    joblib.dump(package, candidate_path)

    old_metrics = evaluate_old_model(texts, labels, free_texts, free_labels)
    report = {
        "dataset": {
            "path": str(dataset_path),
            "rows": len(records),
            "class_distribution": dict(Counter(labels)),
            "unique_product_names": len(set(groups)),
        },
        "old_model": old_metrics,
        "selected_model": metadata,
        "selected_classification_report": best["classification_report"],
        "selected_confusion_matrix": best["confusion_matrix"],
        "free_test_results": best["free_test_results"],
        "candidates": [
            {
                key: value
                for key, value in candidate.items()
                if key not in {"vectorizer", "model", "classification_report", "free_test_results"}
            }
            for candidate in candidates
        ],
    }
    REPORT_PATH.write_text(
        json.dumps(report, indent=2, ensure_ascii=False, default=json_safe),
        encoding="utf-8",
    )

    print(f"Model kandidat: {candidate_path}")
    print(f"Laporan training: {REPORT_PATH}")
    print(json.dumps(metadata, indent=2, ensure_ascii=False, default=json_safe))

    if args.activate:
        archive_dir = BASE_DIR / "models" / "archive"
        archive_dir.mkdir(parents=True, exist_ok=True)
        if ACTIVE_MODEL.exists():
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_path = archive_dir / f"pipeline_decision_tree_umkm_{timestamp}.pkl"
            shutil.copy2(ACTIVE_MODEL, backup_path)
            print(f"Backup model lama: {backup_path}")
        shutil.copy2(candidate_path, ACTIVE_MODEL)
        print(f"Model baru diaktifkan: {ACTIVE_MODEL}")
    else:
        print("Model belum diaktifkan. Jalankan ulang dengan --activate setelah laporan diperiksa.")


if __name__ == "__main__":
    main()
